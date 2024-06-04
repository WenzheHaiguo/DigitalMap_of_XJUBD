# -*- coding: utf-8 -*-

"""
Created on Mon May 27 14:48:06 2024

@author: jinlongshi
@email:  jinlongshi@stu.xju.edu.cn

"""

import urllib.request
from urllib import parse
import json
from openpyxl import load_workbook
from time import sleep

# 常量定义
API_KEY = "dcb33fc98e98d046a94071c2bc1456c4"
CITY = "乌鲁木齐"
EXCEL_FILE_PATH_NAMES = "J:\\Personal file\\Software programmer\\新大博达.xlsx"
EXCEL_SHEET_NAME = "data"
EXCEL_FILE_PATH_DISTANCES = "J:\\Personal file\\Software programmer\\各建筑距离.xlsx"

# 函数定义
def read_excel_data(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    data = [ws[f"A{i+1}"].value for i in range(1, ws.max_row + 1) if ws[f"A{i+1}"].value is not None]
    wb.close()
    return data

def get_encoded_url(base_url, params):
    return parse.quote(base_url.format(**params), safe="/:=&?#+!$,;'@()*[]")

def fetch_json_data(url):
    with urllib.request.urlopen(url) as response:
        return json.loads(response.read())

def get_location(name):
    url = "http://restapi.amap.com/v3/place/text?keywords={}&city={}&output=json&offset=1&page=1&key={}"
    encoded_url = get_encoded_url(url, {"keywords": name, "city": CITY, "key": API_KEY})
    return fetch_json_data(encoded_url)["pois"][0]["location"]

def calculate_distance(origin, destination):
    url = "https://restapi.amap.com/v3/direction/driving?origin={}&destination={}&extensions=all&strategy=5&output=json&key={}"
    encoded_url = get_encoded_url(url, {"origin": origin, "destination": destination, "key": API_KEY})
    return fetch_json_data(encoded_url)["route"]["paths"][0]["distance"]

def write_distances_to_excel(distances, file_path):
    wb = load_workbook(file_path)
    ws = wb["Sheet1"]
    for i, distance in enumerate(distances, start=1):
        ws.cell(row=i, column=3).value = distance
    wb.save(file_path)
    wb.close()

# 主逻辑
def main():
    name_list = read_excel_data(EXCEL_FILE_PATH_NAMES, EXCEL_SHEET_NAME)
    locations = {name: get_location(name) for name in name_list}
    distance_list = []

    for i in range(len(name_list)):
        for j in range(i+1, len(name_list)):
            origin = locations[name_list[i]]
            destination = locations[name_list[j]]
            distance = calculate_distance(origin, destination)
            distance_list.append(distance)
            print(f"Distance between {name_list[i]} and {name_list[j]}: {distance}")

    write_distances_to_excel(distance_list, EXCEL_FILE_PATH_DISTANCES)
    print("Distances have been calculated and saved.")

if __name__ == "__main__":
    main()
